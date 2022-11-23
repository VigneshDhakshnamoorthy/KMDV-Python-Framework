import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class ExcelUtil:
    def __init__(self, file_name):
        try:
            self.wb = load_workbook(f"../Repository/DataRepository/{file_name}.xlsx")
        except FileNotFoundError:
            self.wb = load_workbook(f"{os.getcwd()}/Repository/DataRepository/{file_name}.xlsx")

    def get_test_data(self, text):
        m_ws = self.wb['Main']
        m_row = m_ws.rows
        arr = []
        for mc in m_row:
            if mc[0].value.lower() == text.lower() and mc[1].value.lower() == 'y':

                ws = self.wb[mc[2].value]
                row = ws.rows

                for c in row:

                    if c[0].value == text:
                        row_num = c[0].row
                        row_num += 2
                        while True:
                            if ws[get_column_letter(c[0].column) + str(row_num)].value is not None:
                                if ws[get_column_letter(c[0].column) + str(row_num)].value.lower() == "y":
                                    aer = []
                                    for i in range(0, 7):
                                        aer.append(ws[get_column_letter(c[i].column) + str(row_num)].value)
                                    arr.append(aer)
                            else:
                                break

                            row_num += 1
                        break
        self.wb.close()
        return arr
